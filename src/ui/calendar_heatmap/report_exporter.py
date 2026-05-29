"""Export activity report to Markdown (compact) and Excel formats."""

from __future__ import annotations


def export_markdown(report_data: dict, filepath: str) -> None:
    """Generate a compact Markdown work report.

    Format:
        工作报告：本周 (2026-05-18 ~ 2026-05-24)

        1. #后端
          - 重构认证模块 (30%→80%)：完成JWT验证；重构middleware；开始重构
          - API性能优化 (50%→90%)：优化查询接口；添加缓存层
    """
    period_label = report_data.get("period_label", "")
    date_range = report_data.get("date_range", "")
    tags: dict = report_data.get("tags", {})

    lines = [f"工作报告：{period_label} ({date_range})", ""]

    sorted_tags = sorted(tags.items(), key=lambda kv: -len(kv[1]))
    for idx, (tag_name, tasks) in enumerate(sorted_tags, 1):
        lines.append(f"{idx}. {tag_name}")
        for task in tasks:
            title = task.get("task_title", "未命名任务")
            start_p = task.get("start_progress", 0)
            end_p = task.get("end_progress", 0)
            entries: list[dict] = task.get("entries", [])

            if entries:
                contents = "；".join(e.get("content", "") for e in entries if e.get("content"))
                lines.append(f"  - {title} ({start_p}%→{end_p}%)：{contents}")
            else:
                lines.append(f"  - {title} ({start_p}%→{end_p}%)")
        lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def export_excel(report_data: dict, filepath: str) -> None:
    """Generate an Excel workbook from structured report data."""
    import openpyxl.styles

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作报告"

    headers = ["标签", "任务", "起始进度", "结束进度", "时间", "工作内容"]
    header_fill = openpyxl.styles.PatternFill(start_color="5B8DEF", end_color="5B8DEF", fill_type="solid")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    row = 2
    tags: dict = report_data.get("tags", {})
    for tag_name, tasks in tags.items():
        for task in tasks:
            title = task.get("task_title", "")
            start_p = task.get("start_progress", 0)
            end_p = task.get("end_progress", 0)
            entries: list[dict] = task.get("entries", [])

            if not entries:
                ws.cell(row=row, column=1, value=tag_name)
                ws.cell(row=row, column=2, value=title)
                ws.cell(row=row, column=3, value=f"{start_p}%")
                ws.cell(row=row, column=4, value=f"{end_p}%")
                row += 1
            else:
                for entry in entries:
                    ws.cell(row=row, column=1, value=tag_name)
                    ws.cell(row=row, column=2, value=title)
                    ws.cell(row=row, column=3, value=f"{start_p}%")
                    ws.cell(row=row, column=4, value=f"{end_p}%")
                    ws.cell(row=row, column=5, value=entry.get("ts", ""))
                    ws.cell(row=row, column=6, value=entry.get("content", ""))
                    row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 50

    wb.save(filepath)
