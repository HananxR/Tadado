"""Main window — Todoseq-style layout with custom title bar and adaptive sizing."""

from __future__ import annotations

import ctypes
import datetime as dt
import json
import logging
from ctypes import wintypes
from datetime import date

from PySide6.QtCore import QDateTime, QEvent, QPoint, QSize, Qt, QTime, QTimer
from PySide6.QtGui import QGuiApplication
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QStackedWidget,
    QStatusBar,
    QVBoxLayout,
    QWidget,
)

from ..config import AppConfig
from ..models.repository import TaskRepository
from ..models.task import Task
from ..models.task_filter import TaskFilter
from ..models.task_status import TaskStatus
from ..services.update_checker import UpdateChecker
from ..utils.icon_loader import load_icon
from ..utils.signal_bus import get_signal_bus
from ..utils.widget_utils import combo_width

_log = logging.getLogger("runlog")

from .calendar_heatmap.activity_content_view import ActivityContentView
from .calendar_heatmap.calendar_heatmap_widget import CalendarHeatmapWidget
from .calendar_heatmap.collapse_panel import HeatmapCollapsePanel
from .calendar_heatmap.period_selector import PeriodSelectorBar
from .calendar_heatmap.task_tree_panel import TaskTreePanel
from .controllers.batch_controller import BatchController
from .controllers.filter_coordinator import FilterCoordinator
from .controllers.partition_controller import PartitionController
from .dialogs.about_dialog import AboutDialog
from .dialogs.settings_dialog import SettingsDialog
from .task_list.batch_toolbar import BatchToolbar
from .task_list.task_edit_panel import TaskEditPanel
from .task_list.task_list_model import COL_ARCHIVED, TaskListModel
from .task_list.task_list_view import TaskListView
from .widgets.calendar_popup import CalendarPopup
from .widgets.dropdown import DropdownWidget
from .widgets.filter_bar import FilterBar
from .widgets.progress_dynamics_bar import ProgressDynamicsBar
from .widgets.quick_overview_bar import QuickOverviewBar
from .widgets.status_badge_strip import StatusBadgeStrip
from .widgets.tag_management_panel import TagManagementPanel


class MainWindow(QMainWindow):
    """Desktop task manager with Markdown-first workflow."""

    def __init__(
        self,
        config: AppConfig,
        repository: TaskRepository,
        task_service=None,  # TaskService (optional, for gradual migration)
    ) -> None:
        super().__init__(None, Qt.WindowType.FramelessWindowHint | Qt.WindowType.Window)

        # ── DWM pre-config: must run BEFORE show(), right after HWND creation ──
        # Force native HWND creation so DWM attributes can be set immediately,
        # before DWM ever composites a single frame for this window.
        self.winId()
        from ..utils.win32_theme import (
            enable_window_snap,
            set_window_cloaked,
            set_window_nc_rendering_disabled,
        )
        set_window_nc_rendering_disabled(self)   # never draw native NC buttons
        set_window_cloaked(self, True)           # hide from DWM until fully ready
        enable_window_snap(self)                 # restore WS_THICKFRAME for Aero Snap
        # ──────────────────────────────────────────────────────────────────────

        self.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, True)
        self._config = config
        self._last_applied_theme = config.theme
        self._repository = repository
        self._task_service = task_service
        self._signal_bus = get_signal_bus()
        self._carousel_filter: TaskFilter | None = None
        self._page: int = 0
        self._page_size: int = config.get("general", "page_size", default=20)
        self._total_count: int = 0
        self._current_view: str = "edit"
        self._analysis_date_range: tuple = (None, None)
        self._selection_guard: bool = False  # prevents signal recursion from selectRow()
        self._new_task_sort_active: bool = False  # True after task creation, reset on user nav
        self._setting_sort_internally: bool = False  # guard against self-triggered filter change
        self._update_checker = UpdateChecker(self)

        self.setWindowTitle("Tadado")

        self._setup_custom_title_bar()
        self._setup_status_bar()
        self._setup_central_widget()
        # PartitionController — owns partition lifecycle, replaces _setup_idle_lock + _load_partitions
        self._partition_ctrl = PartitionController(
            self._task_service, self._config,
            self._splitter_stack,
            self._status_partition_btn, self._status_partition_menu,
            self,
        )
        self._partition_ctrl.partition_activated.connect(self._on_partition_activated)
        # BatchController — lazily builds page2
        self._batch_ctrl = BatchController(
            self._task_service, self._config, self._repository,
            self._partition_ctrl, self,
        )
        self._batch_ctrl.data_changed.connect(self._on_data_changed)
        self._batch_ctrl.status_message.connect(self._flash_status)
        # FilterCoordinator — data refresh core for edit view
        self._filter_coordinator = FilterCoordinator(
            self._task_service,
            self._filter_bar,
            self._quick_overview,
            self._task_model,
            self._task_view,
            self._progress_bar,
            self._status_badge,
            self._edit_panel,
            self._status_msg,
            self._config,
            self,
        )
        self._filter_coordinator.status_message.connect(self._flash_status)
        self._connect_signals()
        self._setup_midnight_timer()
        self._partition_ctrl.load_all()
        self._apply_splitter_sizes()

    # ------------------------------------------------------------------
    # Adaptive sizing
    # ------------------------------------------------------------------

    def apply_screen_size(self) -> None:
        self.setMinimumSize(900, 600)
        screen = self.screen() or QGuiApplication.primaryScreen()
        if screen is None:
            self.resize(1050, 680)
            return
        geom = screen.availableGeometry()
        w = min(int(geom.width() * 0.65), 1400)
        h = min(int(geom.height() * 0.72), 900)
        self.resize(w, h)
        self.move(
            (geom.width() - w) // 2 + geom.x(),
            (geom.height() - h) // 2 + geom.y(),
        )
        QTimer.singleShot(100, self._sync_header_alignment)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._apply_splitter_sizes()
        self._apply_batch_splitter_sizes()
        self._sync_header_alignment()

    def _apply_splitter_sizes(self) -> None:
        if self._splitter is None:
            return
        total = self._splitter.width()
        if total > 100:
            self._splitter.setSizes([int(total * 0.50), int(total * 0.50)])

    def _apply_batch_splitter_sizes(self) -> None:
        """Set batch page splitter to 70:30 (existing content : tag panel)."""
        splitter = self._batch_ctrl.splitter
        if splitter is None:
            return
        total = splitter.width()
        if total > 100:
            splitter.setSizes([int(total * 0.80), int(total * 0.20)])

    def _sync_header_alignment(self) -> None:
        """Sync editor header height to match table header for vertical alignment."""
        hh = self._task_view.horizontalHeader()
        if hh:
            h = hh.height()
            if h > 0:
                self._edit_panel.set_header_height(h)

    def refresh_theme(self) -> None:
        self._edit_panel.refresh_theme()
        if hasattr(self, '_analysis_content_view'):
            self._analysis_content_view.refresh_theme()
        if hasattr(self, '_title_icon_btn'):
            self._refresh_title_bar_theme()
        if hasattr(self, '_status_partition_btn'):
            self._refresh_status_partition_style()

    def _refresh_title_bar_theme(self) -> None:
        """Re-apply inline QSS on the title-bar logo button after theme switch."""
        from ..utils.design_tokens import get_tokens as _gt
        t = _gt()
        self._title_icon_btn.setStyleSheet(
            f"QPushButton {{ border: none; background: transparent; padding: 0px; }}"
            f"QPushButton:hover {{ background: {t.accent}20; }}"
        )

    def _refresh_status_partition_style(self) -> None:
        """Re-apply inline QSS on the status-bar partition button after theme switch."""
        from ..utils.design_tokens import get_tokens as _gt
        t = _gt()
        accent = t.accent if t else "#5b8def"
        r, g, b = int(accent[1:3], 16), int(accent[3:5], 16), int(accent[5:7], 16)
        self._status_partition_btn.setStyleSheet(
            f"QPushButton {{ border: none; border-left: 3px solid {accent}; "
            f"background: rgba({r},{g},{b},0.15); border-radius: 4px; "
            f"padding: 2px 8px 2px 6px; font-size: 11px; "
            f"font-weight: bold; color: {accent}; }}"
            f"QPushButton:hover {{ background: rgba({r},{g},{b},0.25); }}"
        )

    # ------------------------------------------------------------------
    # Custom title bar — VS Code style: icon + menu + window buttons
    # ------------------------------------------------------------------

    def _setup_custom_title_bar(self) -> None:
        bar_h = 36

        title_bar = QWidget()
        title_bar.setObjectName("customTitleBar")
        title_bar.setFixedHeight(bar_h)
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(0, 0, 0, 0)
        tb.setSpacing(0)

        # Logo button
        self._title_icon_btn = QPushButton()
        self._title_icon_btn.setIcon(load_icon("app"))
        self._title_icon_btn.setIconSize(QSize(20, 20))
        self._title_icon_btn.setFixedSize(bar_h, bar_h)
        self._title_icon_btn.setFlat(True)
        self._title_icon_btn.setToolTip("返回主界面")
        self._title_icon_btn.clicked.connect(self._on_go_home)
        self._refresh_title_bar_theme()
        tb.addWidget(self._title_icon_btn)

        # Nav buttons (icon + text, flat style) — colors via base.qss
        btn_style = (
            "QPushButton { border: none; background: transparent; padding: 2px 8px; font-size: 11px; }"
        )
        icon_sz = QSize(18, 18)

        nav_items = [
            ("new_task", "新建单任务", self._on_menu_new_draft),
            ("new_multi_task", "新建多任务", self._on_menu_new_multi),
            ("heatmap", "活动分析", lambda: self._switch_view("dashboard")),
            ("task_manage", "任务管理", lambda: self._switch_view("batch")),
            ("settings", "设置", self._on_settings),
        ]
        for icon_name, text, slot in nav_items:
            btn = QPushButton()
            btn.setObjectName("titleBtn")
            btn.setIcon(load_icon(icon_name))
            btn.setIconSize(icon_sz)
            btn.setText(text)
            btn.setFlat(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(btn_style)
            btn.clicked.connect(slot)
            tb.addWidget(btn)

        # Help button with dropdown
        help_btn = QPushButton()
        help_btn.setIcon(load_icon("help"))
        help_btn.setIconSize(icon_sz)
        help_btn.setText("帮助")
        help_btn.setFlat(True)
        help_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        help_btn.setStyleSheet(btn_style)
        help_menu = QMenu(help_btn)
        help_menu.addAction("帮助文档(&D)", self._on_help_docs)
        help_menu.addSeparator()
        help_menu.addAction("关于(&A)", self._on_about)
        help_btn.setMenu(help_menu)
        help_btn.clicked.connect(lambda: help_btn.showMenu())
        tb.addWidget(help_btn)

        # Store the right edge of nav buttons for hit-test (logo 36 + ~110px per button * 6 + help ~80px)
        self._title_nav_right = 36 + 110 * 6 + 80

        tb.addStretch()

        # Right-side window buttons (icon only) — colors via base.qss
        right_btn_style = (
            "QPushButton { border: none; background: transparent; padding: 0px; }"
        )
        right_items = [
            ("tray_hide", "缩小到托盘", self.hide),
            ("window_minimize", "最小化", self._on_minimize),
            ("fullscreen_toggle", "切换全屏", self._toggle_fullscreen),
            ("window_close", "关闭", self.close),
        ]
        for icon_name, tip, slot in right_items:
            btn = QPushButton()
            btn.setIcon(load_icon(icon_name))
            btn.setIconSize(QSize(20, 20))
            btn.setFixedSize(bar_h, bar_h)
            btn.setFlat(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setToolTip(tip)
            btn.setStyleSheet(right_btn_style)
            btn.clicked.connect(slot)
            tb.addWidget(btn)

        self.setMenuWidget(title_bar)

    def _toggle_fullscreen(self) -> None:
        if self.isFullScreen():
            self.showNormal()
            self.apply_screen_size()
        else:
            self.showFullScreen()

    # ------------------------------------------------------------------
    # Win32 native event — window resize + title-bar drag
    # ------------------------------------------------------------------

    def nativeEvent(self, event_type, message):
        if event_type == b"windows_generic_MSG":
            msg = wintypes.MSG.from_address(message.__int__())
            if msg.message == 0x0084:  # WM_NCHITTEST
                return self._nc_hit_test(msg)
            elif msg.message == 0x0083:  # WM_NCCALCSIZE
                # Extend the client area to cover the entire window rect.
                # wParam==0 → simple RECT; wParam==1 → NCCALCSIZE_PARAMS.
                # In both cases we return True,0 to claim we handled it and
                # request the full window area — the invisible border added
                # by WS_THICKFRAME must not shrink our client area.
                return True, 0
            elif msg.message == 0x0024:  # WM_GETMINMAXINFO
                # Let DefWindowProc handle it; the default maximised monitor
                # rect works correctly with WS_THICKFRAME on Win10/11.
                return False, 0
        return super().nativeEvent(event_type, message)

    def _nc_hit_test(self, msg) -> tuple:
        raw_low = msg.lParam & 0xFFFF
        raw_high = (msg.lParam >> 16) & 0xFFFF
        x = ctypes.c_short(raw_low).value
        y = ctypes.c_short(raw_high).value
        dpr = self.devicePixelRatioF()
        if dpr != 1.0:
            x = int(x / dpr)
            y = int(y / dpr)
        border = 8  # match Win10/11 standard invisible resize border
        g = self.geometry()
        title_h = 36
        if g.y() <= y < g.y() + title_h:
            # Use childAt() to distinguish buttons from empty draggable space.
            # Any child widget under the cursor → HTCLIENT (button works).
            # Empty area → HTCAPTION (entire title bar can drag to Snap).
            bar = self.menuWidget()
            if bar is not None:
                local = bar.mapFromGlobal(QPoint(x, y))
                if bar.childAt(local) is not None:
                    return False, 0  # HTCLIENT — button receives click
            return True, 2  # HTCAPTION — draggable
        left = x < g.x() + border
        right = x > g.x() + g.width() - border
        top = y < g.y() + border
        bottom = y > g.y() + g.height() - border
        if top and left:
            return True, 13
        if top and right:
            return True, 14
        if bottom and left:
            return True, 16
        if bottom and right:
            return True, 17
        if left:
            return True, 10
        if right:
            return True, 11
        if bottom:
            return True, 15
        return False, 0

    def _on_minimize(self) -> None:
        """Minimize button — respects *minimize_to_tray* config.

        When enabled, hide directly to tray (no taskbar flash).
        Otherwise do a normal minimize to the taskbar.
        """
        if self._config.minimize_to_tray:
            self.hide()
        else:
            self.showMinimized()

    def changeEvent(self, event) -> None:
        """Intercept minimize: hide to tray when *minimize_to_tray* is enabled."""
        if (
            event.type() == QEvent.Type.WindowStateChange
            and self.windowState() & Qt.WindowState.WindowMinimized
        ):
            if self._config.minimize_to_tray:
                self.hide()
                event.ignore()
                return
        super().changeEvent(event)

    # ------------------------------------------------------------------
    # Tool bar
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Central widget
    # ------------------------------------------------------------------

    def _setup_central_widget(self) -> None:
        self._stack = QStackedWidget()

        # === Page 0: Task view ===
        task_page = QWidget()
        task_layout = QVBoxLayout(task_page)
        task_layout.setContentsMargins(8, 4, 8, 4)
        task_layout.setSpacing(2)

        # Row 1: QuickOverviewBar only (presets + carousel)
        self._top_bar = QWidget()
        top_bar_layout = QHBoxLayout(self._top_bar)
        top_bar_layout.setContentsMargins(4, 0, 4, 0)
        top_bar_layout.setSpacing(0)

        self._quick_overview = QuickOverviewBar(self._repository, max_items=2, group_size=2, interval_seconds=5)
        self._quick_overview.preset_activated.connect(self._on_quick_preset)
        self._quick_overview.task_clicked.connect(self._on_carousel_clicked)
        top_bar_layout.addWidget(self._quick_overview, 1)
        task_layout.addWidget(self._top_bar)

        # Heatmap widget (created here, used in heatmap page)
        self._heatmap_widget = CalendarHeatmapWidget(self._repository, self._config)
        self._heatmap_widget.back_requested.connect(lambda: self._switch_view("edit"))

        # Row 2: FilterBar + StatusBadgeStrip (same row, StatusBadgeStrip right-aligned)
        filter_row = QWidget()
        filter_row_layout = QHBoxLayout(filter_row)
        filter_row_layout.setContentsMargins(4, 0, 4, 0)
        filter_row_layout.setSpacing(6)

        self._filter_bar = FilterBar()
        self._filter_bar.set_sort(self._config.default_sort)
        filter_row_layout.addWidget(self._filter_bar, 1)

        self._status_badge = StatusBadgeStrip(self._repository)
        self._status_badge.filter_changed.connect(self._on_filter_changed)
        filter_row_layout.addWidget(self._status_badge)
        task_layout.addWidget(filter_row)

        # Splitter: task list (left) + edit panel (right)
        from PySide6.QtWidgets import QStackedLayout as _QStackedLayout
        self._splitter_container = QWidget()
        self._splitter_stack = _QStackedLayout(self._splitter_container)
        self._splitter_stack.setContentsMargins(0, 0, 0, 0)
        self._splitter_stack.setStackingMode(_QStackedLayout.StackingMode.StackOne)

        self._splitter = QSplitter(Qt.Orientation.Horizontal)
        self._splitter.setHandleWidth(2)
        self._splitter.setChildrenCollapsible(False)
        self._splitter.setStretchFactor(0, 1)
        self._splitter.setStretchFactor(1, 1)

        # === Left panel: BatchToolbar + TaskListView + Pagination ===
        left_panel = QWidget()
        left_panel.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 2, 0)
        left_layout.setSpacing(2)

        self._batch_toolbar = BatchToolbar()
        left_layout.addWidget(self._batch_toolbar)

        self._task_model = TaskListModel()
        self._task_view = TaskListView(
            self._repository, task_service=self._task_service,
        )
        self._task_view.set_model(self._task_model)
        self._task_view.setColumnHidden(COL_ARCHIVED, True)  # 归档列仅管理视图可见
        self._task_view.task_selected.connect(self._on_view_task_selected)
        self._task_view.detail_requested.connect(self._on_detail_requested)
        # Batch operations from right-click menu
        self._task_view.batch_status_change.connect(self._on_batch_status_change)
        self._task_view.batch_urgency_change.connect(self._on_batch_urgency_change)
        self._task_view.batch_delete.connect(self._on_batch_delete)
        self._task_view.batch_suspend.connect(self._on_batch_suspend)
        self._task_view.batch_restart.connect(self._on_batch_restart)
        self._task_view.batch_postpone.connect(self._on_batch_postpone)
        self._task_view.batch_move_partition.connect(self._on_batch_move_partition)
        left_layout.addWidget(self._task_view, 1)

        # Pagination
        page_widget = QWidget()
        page_row = QHBoxLayout(page_widget)
        page_row.setContentsMargins(4, 2, 4, 2)
        page_row.setSpacing(4)
        page_row.addStretch()
        self._prev_page_btn = QPushButton("‹")
        self._prev_page_btn.setObjectName("navBtn")
        self._prev_page_btn.setFixedWidth(28)
        self._prev_page_btn.clicked.connect(self._on_page_prev)
        page_row.addWidget(self._prev_page_btn)
        self._page_label = QLabel("1 / 1")
        self._page_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        page_row.addWidget(self._page_label)
        self._next_page_btn = QPushButton("›")
        self._next_page_btn.setObjectName("navBtn")
        self._next_page_btn.setFixedWidth(28)
        self._next_page_btn.clicked.connect(self._on_page_next)
        page_row.addWidget(self._next_page_btn)
        self._page_size_combo = DropdownWidget()
        self._page_size_combo.setFixedWidth(combo_width(4))
        for n in ["20", "50", "100"]:
            self._page_size_combo.addItem(n, int(n))
        self._page_size_combo.setCurrentText(str(self._page_size))
        self._page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)
        page_row.addWidget(self._page_size_combo)
        left_layout.addWidget(page_widget)

        self._splitter.addWidget(left_panel)

        # === Right panel: ProgressDynamicsBar + TaskEditPanel ===
        right_panel = QWidget()
        right_panel.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(2, 0, 0, 0)
        right_layout.setSpacing(2)

        self._progress_bar = ProgressDynamicsBar(self._repository)
        right_layout.addWidget(self._progress_bar)
        self._progress_bar.progress_filter_activated.connect(self._on_progress_filter)
        self._progress_bar.task_clicked.connect(self._on_carousel_clicked)

        self._edit_panel = TaskEditPanel(
            self._repository, self._task_model,
            task_service=self._task_service,
        )
        right_layout.addWidget(self._edit_panel, 1)
        self._splitter.addWidget(right_panel)

        self._splitter_stack.addWidget(self._splitter)
        # Password mask overlay
        self._partition_mask = QWidget()
        self._partition_mask.setObjectName("partitionMask")
        mask_layout = QVBoxLayout(self._partition_mask)
        mask_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mask_hint = QLabel("此分区已加密\n请输入密码查看内容")
        mask_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mask_hint.setStyleSheet(
            "QLabel { font-size: 16px; font-weight: bold;"
            " background: transparent; border: none; }"
        )
        mask_layout.addWidget(mask_hint)
        unlock_btn = QPushButton("输入密码解锁")
        unlock_btn.setObjectName("saveBtn")
        unlock_btn.setFixedWidth(140)
        # Deferred connect: _partition_ctrl created after _setup_central_widget
        unlock_btn.clicked.connect(lambda: self._partition_ctrl.unlock())
        mask_btn_row = QHBoxLayout()
        mask_btn_row.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mask_btn_row.addWidget(unlock_btn)
        mask_layout.addLayout(mask_btn_row)
        self._splitter_stack.addWidget(self._partition_mask)
        self._splitter_stack.setCurrentIndex(0)
        task_layout.addWidget(self._splitter_container, 1)

        self._stack.addWidget(task_page)

        # Page 1 & 2 are built lazily on first access
        self._stack.insertWidget(1, QWidget())  # placeholder for page 1
        self._stack.insertWidget(2, QWidget())  # placeholder for page 2
        self._page1_built = False
        self._page2_built = False

        self._batch_page = 0
        self._batch_total_count = 0
        self._batch_pending_action: dict = {}

        self.setCentralWidget(self._stack)

    # ------------------------------------------------------------------
    # Lazy page builders
    # ------------------------------------------------------------------

    def _build_page1(self) -> None:
        """Build Activity Analysis page on first access."""
        from .calendar_heatmap.heatmap_stats_panel import HeatmapStatsPanel

        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(8)

        heatmap_label = QLabel("活动热力图")
        heatmap_label.setObjectName("analysisSectionLabel")
        layout.addWidget(heatmap_label)

        ht_row = QWidget()
        ht_layout = QHBoxLayout(ht_row)
        ht_layout.setContentsMargins(0, 0, 0, 0)
        ht_layout.addWidget(self._heatmap_widget.nav_bar)
        ht_layout.addStretch()
        self._analysis_stats = HeatmapStatsPanel()
        self._analysis_stats.setFixedHeight(28)
        ht_layout.addWidget(self._analysis_stats)
        layout.addWidget(ht_row)

        collapsible = HeatmapCollapsePanel(self._heatmap_widget)
        layout.addWidget(collapsible, 0)

        report_label = QLabel("活动报告")
        report_label.setObjectName("analysisSectionLabel")
        layout.addWidget(report_label)

        period_row = QWidget()
        period_layout = QHBoxLayout(period_row)
        period_layout.setContentsMargins(0, 4, 0, 4)
        period_layout.setSpacing(6)

        self._analysis_period_selector = PeriodSelectorBar()
        self._analysis_period_selector.period_changed.connect(self._on_analysis_period_changed)
        period_layout.addWidget(self._analysis_period_selector, 1)

        self._analysis_search = QLineEdit()
        self._analysis_search.setPlaceholderText("搜索活动内容...")
        self._analysis_search.setFixedWidth(150)
        self._analysis_search.setFixedHeight(28)
        self._analysis_search.textChanged.connect(self._on_analysis_search_changed)
        period_layout.addWidget(self._analysis_search)

        export_btn = QPushButton("导出")
        export_btn.setObjectName("exportBtn")
        export_btn.setFixedHeight(28)
        export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        export_menu = QMenu(export_btn)
        export_menu.addAction("导出 Markdown", self._on_export_analysis_md)
        export_menu.addAction("导出 Excel", self._on_export_analysis_xlsx)
        export_menu.addAction("导出 TXT", self._on_export_analysis_txt)
        export_btn.setMenu(export_menu)
        export_btn.clicked.connect(lambda: export_btn.showMenu())
        period_layout.addWidget(export_btn)
        layout.addWidget(period_row)

        self._analysis_splitter = QSplitter(Qt.Orientation.Horizontal)
        self._analysis_splitter.setHandleWidth(1)
        self._analysis_splitter.setChildrenCollapsible(False)

        self._analysis_task_tree = TaskTreePanel(self._repository)
        self._analysis_task_tree.tag_selected.connect(self._on_analysis_tag_selected)
        self._analysis_splitter.addWidget(self._analysis_task_tree)

        self._analysis_content_view = ActivityContentView()
        self._analysis_content_view.prev_requested.connect(self._on_analysis_prev)
        self._analysis_content_view.next_requested.connect(self._on_analysis_next)
        self._analysis_splitter.addWidget(self._analysis_content_view)

        self._analysis_splitter.setStretchFactor(0, 1)
        self._analysis_splitter.setStretchFactor(1, 3)
        layout.addWidget(self._analysis_splitter, 1)

        self._heatmap_widget.grid.date_clicked.connect(self._on_heatmap_date_clicked)

        old = self._stack.widget(1)
        self._stack.removeWidget(old)
        if old:
            old.deleteLater()
        self._stack.insertWidget(1, page)
        self._page1_built = True

    def _build_page2(self) -> None:
        """Build Task Management Console page — delegated to BatchController."""
        page = self._batch_ctrl.build_page()
        old = self._stack.widget(2)
        self._stack.removeWidget(old)
        if old:
            old.deleteLater()
        self._stack.insertWidget(2, page)
        self._page2_built = True

    def _on_new_multi_task(self) -> None:
        if not self.isVisible() and self._edit_panel.has_unsaved_draft():
            self._edit_panel.discard_draft()
        elif not self._guard_draft():
            return
        if self._splitter_stack.currentIndex() == 1:
            if self._partition_ctrl.passwords.get(self._partition_ctrl.active_id, ""):
                self._partition_ctrl.unlock()
                if self._splitter_stack.currentIndex() == 1:
                    return
            else:
                self._splitter_stack.setCurrentIndex(0)
        self._stack.setCurrentIndex(0)
        self.show()
        self.setWindowState(self.windowState() & ~Qt.WindowState.WindowMinimized)
        self.raise_()
        self.activateWindow()
        self._edit_panel.set_active_partition(self._partition_ctrl.active_id)
        self._edit_panel.create_draft_multi()
        self._apply_splitter_sizes()

    # ------------------------------------------------------------------
    # Status bar
    # ------------------------------------------------------------------

    def _setup_status_bar(self) -> None:
        self._status_bar = QStatusBar()
        self._status_bar.setSizeGripEnabled(True)

        # Partition selector button (prominent, left side)
        self._status_partition_btn = QPushButton("● 切换分区")
        self._status_partition_btn.setFlat(True)
        self._status_partition_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._refresh_status_partition_style()
        self._status_partition_menu = QMenu(self._status_partition_btn)
        self._status_partition_btn.setMenu(self._status_partition_menu)
        self._status_partition_btn.clicked.connect(lambda: self._status_partition_btn.showMenu())
        self._status_bar.addWidget(self._status_partition_btn)

        # Stats + motd text
        self._status_msg = QLabel("就绪")
        self._status_bar.addWidget(self._status_msg, 1)

        # Right: clock
        self._status_clock = QLabel()
        self._status_clock.setStyleSheet("QLabel { margin-right: 4px; }")
        self._status_bar.addPermanentWidget(self._status_clock)
        self._update_status_clock()
        self._clock_timer = QTimer(self)
        self._clock_timer.timeout.connect(self._update_status_clock)
        self._clock_timer.start(1000)
        self.setStatusBar(self._status_bar)

    def _update_status_clock(self) -> None:
        self._status_clock.setText(dt.datetime.now().strftime("%Y年%m月%d日 %I:%M:%S %p"))

    # ------------------------------------------------------------------
    # Signals
    # ------------------------------------------------------------------

    def _connect_signals(self) -> None:
        bus = self._signal_bus
        fc = self._filter_coordinator

        # Data refresh → FilterCoordinator (ignore signal args to keep defaults)
        bus.scan_completed.connect(lambda *_: fc.refresh())
        bus.task_updated.connect(lambda *_: fc.refresh())
        bus.task_deleted.connect(lambda *_: fc.refresh())
        bus.task_status_changed.connect(lambda *_: fc.refresh())
        bus.batch_operation_completed.connect(lambda *_: fc.refresh())
        bus.archive_completed.connect(lambda *_: fc.refresh())
        bus.tag_changed.connect(lambda *_: fc.refresh())

        # Task creation (special sort) → FilterCoordinator
        bus.task_created.connect(fc.handle_new_task_sort)
        bus.tasks_bulk_created.connect(fc.handle_tasks_bulk_created)

        # Partitions → PartitionController
        bus.partitions_changed.connect(self._partition_ctrl.load_all)

        # Config → MainWindow (cross-cutting)
        bus.config_changed.connect(self._on_config_changed)

        # Heatmap refresh (widget concern)
        bus.task_created.connect(self._on_heatmap_data_changed)
        bus.task_updated.connect(self._on_heatmap_data_changed)
        bus.task_deleted.connect(self._on_heatmap_data_changed)
        bus.task_status_changed.connect(self._on_heatmap_data_changed)
        bus.batch_operation_completed.connect(self._on_heatmap_data_changed)

        # Edit-view toolbar (still in MainWindow for now)
        self._batch_toolbar.select_all_requested.connect(
            lambda: self._task_model.set_checked_ids(
                set(t.id for t in self._task_model.tasks)
            )
        )
        self._batch_toolbar.deselect_all_requested.connect(
            lambda: self._task_model.set_checked_ids(set())
        )

    # ------------------------------------------------------------------
    # Slots: data refresh
    # ------------------------------------------------------------------

    def _on_data_changed(self, *args) -> None:
        if hasattr(self, '_splitter_stack') and self._splitter_stack.currentIndex() == 1:
            return
        if hasattr(self, '_progress_bar'):
            self._progress_bar.reset_to_unclicked()
        _sizes = self._splitter.sizes() if hasattr(self, '_splitter') and self._splitter else None
        f = self._build_filter_with_sort()
        self._refresh_all_views(f, reset_page=False)
        if _sizes:
            self._splitter.setSizes(_sizes)
        # Re-select task if triggered by a task update signal (keep current task open)
        if args and hasattr(args[0], 'id'):
            self._select_and_load_task(args[0].id)

    def _on_tasks_bulk_created(self, count: int, task_ids: list) -> None:
        """Handle multi-task creation: switch to creation-time sort, refresh, highlight first."""
        self._new_task_sort_active = True
        self._filter_bar.blockSignals(True)
        self._filter_bar.set_sort("created")
        self._filter_bar.reset()
        self._filter_bar._debounce.stop()  # 杀死 reset() 残留的 300ms debounce，避免竞态
        self._filter_bar.blockSignals(False)
        if hasattr(self, '_quick_overview') and self._quick_overview.active_preset != "today":
            self._new_task_sort_active = False  # 临时清除，避免 _on_quick_preset 恢复默认排序
            self._quick_overview.activate_preset("today")
            self._new_task_sort_active = True
        self._on_data_changed()
        if self._task_model.tasks:
            self._on_task_selected(self._task_model.tasks[0])

    def _build_filter_with_sort(self) -> TaskFilter:
        """Build filter with FilterBar's sort as base, overlay scope from quick-overview + partition."""
        f = self._filter_bar.build_filter()  # preserves sort + search + urgencies
        # 从速览栏当前预设直接获取时间窗口和状态过滤（不受 _carousel_filter 覆写影响）
        if hasattr(self, '_quick_overview'):
            overview_f = self._quick_overview.build_filter()
            f.created_to = overview_f.created_to
            f.statuses = overview_f.statuses
        if self._carousel_filter is not None:
            f.date_from = self._carousel_filter.date_from
            f.date_to = self._carousel_filter.date_to
            f.activity_field = self._carousel_filter.activity_field
            f.activity_min = self._carousel_filter.activity_min
            f.partition_id = self._carousel_filter.partition_id or self._partition_ctrl.active_id or None  # "" → None
        else:
            f.partition_id = self._partition_ctrl.active_id or None  # "" → None
        return f

    def _refresh_all_views(self, filter_: TaskFilter, reset_page: bool = True) -> None:
        if reset_page:
            self._reset_pagination()
        filter_.partition_id = filter_.partition_id or self._partition_ctrl.active_id or None  # "" → None
        all_tasks, self._total_count = self._task_service.search_with_total(filter_)
        # Paginate table display — full list still passed to overview / progress bar
        start = self._page * self._page_size
        page_tasks = all_tasks[start:start + self._page_size]
        self._task_model.set_offset(start)
        self._task_model.load_tasks(page_tasks)
        self._quick_overview.set_items(all_tasks)
        self._update_page_label()
        self._update_status_bar(filter_)
        self._status_badge.refresh(filter_.date_from, filter_.date_to)
        self._progress_bar.set_items(all_tasks)

    def _on_task_created(self, task) -> None:
        self._new_task_sort_active = True
        self._filter_bar.blockSignals(True)
        self._filter_bar.set_sort("created")
        self._filter_bar.reset()
        self._filter_bar._debounce.stop()  # 杀死 reset() 残留的 300ms debounce，避免竞态
        self._filter_bar.blockSignals(False)
        if hasattr(self, '_quick_overview') and self._quick_overview.active_preset != "today":
            self._new_task_sort_active = False  # 临时清除，避免 _on_quick_preset 恢复默认排序
            self._quick_overview.activate_preset("today")
            self._new_task_sort_active = True
        self._on_data_changed()
        self._on_task_selected(task)

    def _select_and_load_task(self, task_id: str) -> None:
        """Find task by ID and delegate to _on_task_selected (unified凸显 entry)."""
        for row in range(self._task_model.rowCount()):
            if self._task_model.tasks[row].id == task_id:
                self._on_task_selected(self._task_model.tasks[row])
                return

    def _on_task_deleted(self, task_id: str) -> None:
        self._on_data_changed()

    def _on_batch_completed(self) -> None:
        self._on_data_changed()

    # ------------------------------------------------------------------
    # Batch page methods
    # ------------------------------------------------------------------

    def _refresh_batch_page(self) -> None:
        """Refresh the task management page applying all sidebar filters."""
        if not hasattr(self, '_batch_task_model'):
            return
        f = TaskFilter()
        f.sort_by = self._filter_bar.build_filter().sort_by  # inherit main sort
        f.partition_id = self._partition_ctrl.active_id
        f.search_text = self._batch_search.text().strip()
        # Status
        sd = self._batch_status_combo.currentData()
        if sd is not None:
            f.statuses = {sd}
        # Priority / Urgency
        pd = self._batch_priority_combo.currentData()
        if pd is not None:
            f.urgencies = {pd}
        # Created time
        f.created_from = self._read_date_edit('_batch_created_from')
        f.created_to = self._read_date_edit('_batch_created_to')
        # Deadline time
        f.date_from = self._read_date_edit('_batch_deadline_from')
        f.date_to = self._read_date_edit('_batch_deadline_to')
        # Progress
        lo, hi = self._batch_progress_combo.currentData()
        f.progress_min = lo
        f.progress_max = hi
        # Tags (strip leading # for consistency with UI display format)
        tag_text = self._batch_tag_input.text().strip()
        if tag_text:
            f.tags = set(t.strip().lstrip("#").strip() for t in tag_text.split() if t.strip())
        # Archive status
        arc = self._batch_archive_combo.currentData()
        if arc == "all" or arc == "archived":
            f.show_archived = True

        if arc == "archived":
            # Load all tasks (no limit), filter client-side, then paginate manually
            f.limit = None
            f.offset = 0
            tasks = [t for t in self._task_service.search(f) if t.archived]
            self._batch_total_count = len(tasks)
            start = self._batch_page * self._batch_page_size
            tasks = tasks[start:start + self._batch_page_size]
        else:
            f.limit = self._batch_page_size
            f.offset = self._batch_page * self._batch_page_size
            tasks, self._batch_total_count = self._task_service.search_with_total(f)
        self._batch_task_model.set_offset(self._batch_page * self._batch_page_size)
        self._batch_task_model.load_tasks(tasks)
        self._update_batch_pagination()

    def _read_date_edit(self, attr: str) -> date | None:
        """Parse yyyy-MM-dd from a QLineEdit attribute, return date or None."""
        le = getattr(self, attr, None)
        if le is None:
            return None
        txt = le.text().strip()
        if not txt:
            return None
        try:
            return date.fromisoformat(txt)
        except ValueError:
            return None

    def _open_date_popup(self, line_edit: QLineEdit) -> None:
        """Open CalendarPopup and set result into the QLineEdit."""
        txt = line_edit.text().strip()
        initial = date.fromisoformat(txt) if txt else date.today()
        popup = CalendarPopup(initial, self)
        popup.date_selected.connect(lambda qd: (
            line_edit.setText(qd.toPython().isoformat()),
            self._on_batch_filter_changed()
        ))
        popup.smart_place(line_edit)
        popup.exec()

    def _update_batch_pagination(self) -> None:
        if self._batch_page_size <= 0:
            self._batch_page_label.setText("全部")
            return
        total_pages = max(1, (self._batch_total_count + self._batch_page_size - 1) // self._batch_page_size)
        self._batch_page_label.setText(f"{self._batch_page + 1} / {total_pages}")
        self._batch_prev_btn.setEnabled(self._batch_page > 0)
        self._batch_next_btn.setEnabled(self._batch_page < total_pages - 1)

    def _on_batch_search(self) -> None:
        self._batch_page = 0
        self._refresh_batch_page()

    def _on_batch_filter_changed(self) -> None:
        self._batch_page = 0
        self._refresh_batch_page()

    def _on_batch_page_prev(self) -> None:
        if self._batch_page > 0:
            self._batch_page -= 1
            self._refresh_batch_page()
            if hasattr(self, '_batch_task_model') and self._batch_task_model.rowCount() > 0:
                self._batch_task_model.set_highlighted_task(
                    self._batch_task_model.tasks[0].id)

    def _on_batch_page_next(self) -> None:
        total_pages = max(1, (self._batch_total_count + self._batch_page_size - 1) // self._batch_page_size)
        if self._batch_page < total_pages - 1:
            self._batch_page += 1
            self._refresh_batch_page()
            if hasattr(self, '_batch_task_model') and self._batch_task_model.rowCount() > 0:
                self._batch_task_model.set_highlighted_task(
                    self._batch_task_model.tasks[0].id)

    def _on_batch_page_size_changed(self, index: int) -> None:
        widget = self.sender()
        if widget:
            self._batch_page_size = widget.itemData(index)
            self._batch_page = 0
            self._refresh_batch_page()

    def _on_edit_select_all(self) -> None:
        if hasattr(self, '_task_model'):
            ids = set(t.id for t in self._task_model.tasks)
            self._task_model.set_checked_ids(ids)

    def _on_edit_deselect_all(self) -> None:
        if hasattr(self, '_task_model'):
            self._task_model.set_checked_ids(set())

    def _on_batch_select_all(self) -> None:
        if hasattr(self, '_batch_task_model'):
            ids = set(t.id for t in self._batch_task_model.tasks)
            self._batch_task_model.set_checked_ids(ids)

    def _on_batch_deselect_all(self) -> None:
        if hasattr(self, '_batch_task_model'):
            self._batch_task_model.set_checked_ids(set())

    def _on_batch_task_selected(self, task: Task) -> None:
        """Highlight the selected task in the batch view."""
        self._batch_task_model.set_highlighted_task(task.id)

    def _on_batch_model_data_changed(self) -> None:
        if hasattr(self, '_batch_toolbar2'):
            ids = self._batch_task_model.checked_task_ids()
            self._batch_toolbar2.set_selected(ids)

    def _on_batch_status_change(self, ids: list[str], status) -> None:
        if self._current_view == "edit":
            reply = QMessageBox.question(
                self, "确认操作",
                f"确认更改 {len(ids)} 个任务的状态？",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Ok:
                self._task_service.batch_update_status(ids, status)
                self._task_model.set_checked_ids(set())
                self._on_data_changed()
                self._batch_toolbar.reset_toggle()
                self._flash_status(f"已更改 {len(ids)} 个任务状态")
        else:
            self._confirm_label.setText(f"确认更改 {len(ids)} 个任务的状态？")
            self._confirm_bar.setVisible(True)
            self._batch_pending_action = {"action": "status", "ids": ids, "status": status}
            self._confirm_ok_btn.clicked.disconnect()
            self._confirm_ok_btn.clicked.connect(self._execute_batch_status)

    def _execute_batch_status(self) -> None:
        action = self._batch_pending_action
        self._task_service.batch_update_status(action["ids"], action["status"])
        self._hide_confirm()
        self._refresh_batch_page()
        self._on_data_changed()
        self._flash_status(f"已更改 {len(action['ids'])} 个任务状态")

    def _on_batch_urgency_change(self, ids: list[str], urgency: int) -> None:
        """Handle batch urgency change from toolbar."""
        self._task_service.batch_update_urgency(ids, urgency)
        if self._current_view == "batch":
            self._batch_task_model.deselect_all()
            self._batch_toolbar2.reset_toggle()
            self._refresh_batch_page()
        else:
            self._task_model.set_checked_ids(set())
            self._batch_toolbar.reset_toggle()
            current = self._edit_panel.current_task()
            if current and current.id in ids:
                updated = self._task_service.get_task(current.id)
                if updated:
                    self._edit_panel.load_task(updated)
        self._on_data_changed()
        self._flash_status(f"已更改 {len(ids)} 个任务优先级")

    def _on_batch_delete(self, ids: list[str]) -> None:
        if self._current_view == "edit":
            reply = QMessageBox.question(
                self, "确认删除",
                f"确认删除 {len(ids)} 个任务？此操作不可撤销。",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Ok:
                self._task_service.batch_delete(ids)
                self._task_model.set_checked_ids(set())
                self._on_data_changed()
                self._batch_toolbar.reset_toggle()
                self._flash_status(f"已删除 {len(ids)} 个任务")
        else:
            self._confirm_label.setText(f"⚠ 确认删除 {len(ids)} 个任务？此操作不可撤销。")
            self._confirm_bar.setVisible(True)
            self._batch_pending_action = {"action": "delete", "ids": ids}
            self._confirm_ok_btn.clicked.disconnect()
            self._confirm_ok_btn.clicked.connect(self._execute_batch_delete)

    def _execute_batch_delete(self) -> None:
        action = self._batch_pending_action
        self._task_service.batch_delete(action["ids"])
        self._hide_confirm()
        self._refresh_batch_page()
        self._on_data_changed()
        self._flash_status(f"已删除 {len(action['ids'])} 个任务")

    def _on_batch_suspend(self, ids: list[str]) -> None:
        if self._current_view == "edit":
            reply = QMessageBox.question(
                self, "确认操作",
                f"确认中止 {len(ids)} 个任务？",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Ok:
                self._task_service.batch_suspend(ids)
                self._task_model.set_checked_ids(set())
                self._on_data_changed()
                self._batch_toolbar.reset_toggle()
                self._flash_status(f"已中止 {len(ids)} 个任务")
        else:
            self._confirm_label.setText(f"确认中止 {len(ids)} 个任务？")
            self._confirm_bar.setVisible(True)
            self._batch_pending_action = {"action": "suspend", "ids": ids}
            self._confirm_ok_btn.clicked.disconnect()
            self._confirm_ok_btn.clicked.connect(self._execute_batch_suspend)

    def _execute_batch_suspend(self) -> None:
        action = self._batch_pending_action
        self._task_service.batch_suspend(action["ids"])
        self._hide_confirm()
        self._refresh_batch_page()
        self._on_data_changed()
        self._flash_status(f"已中止 {len(action['ids'])} 个任务")

    def _on_batch_restart(self, ids: list[str]) -> None:
        if self._current_view == "edit":
            reply = QMessageBox.question(
                self, "确认操作",
                f"确认重启 {len(ids)} 个任务？",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Ok:
                self._task_service.batch_restart(ids)
                self._task_model.set_checked_ids(set())
                self._on_data_changed()
                self._batch_toolbar.reset_toggle()
                self._flash_status(f"已重启 {len(ids)} 个任务")
        else:
            self._confirm_label.setText(f"确认重启 {len(ids)} 个任务？")
            self._confirm_bar.setVisible(True)
            self._batch_pending_action = {"action": "restart", "ids": ids}
            self._confirm_ok_btn.clicked.disconnect()
            self._confirm_ok_btn.clicked.connect(self._execute_batch_restart)

    def _execute_batch_restart(self) -> None:
        action = self._batch_pending_action
        self._task_service.batch_restart(action["ids"])
        self._hide_confirm()
        self._refresh_batch_page()
        self._on_data_changed()
        self._flash_status(f"已重启 {len(action['ids'])} 个任务")

    def _on_batch_postpone(self, ids: list[str], days: int) -> None:
        reply = QMessageBox.question(
            self, "确认操作",
            f"确认将 {len(ids)} 个任务的截止时间延后 {days} 天？",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
        )
        if reply == QMessageBox.StandardButton.Ok:
            self._task_service.batch_postpone(ids, days)
            if self._current_view == "edit":
                self._task_model.set_checked_ids(set())
                self._batch_toolbar.reset_toggle()
            else:
                self._refresh_batch_page()
            self._on_data_changed()
            self._flash_status(f"已延后 {len(ids)} 个任务")

    def _on_batch_move_partition(self, ids: list[str]) -> None:
        """Move selected tasks to a different partition with password checks."""
        from PySide6.QtWidgets import (
            QDialog,
            QDialogButtonBox,
            QListWidget,
            QListWidgetItem,
            QVBoxLayout,
        )

        from_partition_id = self._partition_ctrl.active_id or ""
        name_map = self._task_service.get_partition_name_map()
        from_name = name_map.get(from_partition_id, "未分配") if from_partition_id else "未分配"

        # ── Step 1: Verify FROM partition password ──
        from_pw = self._partition_ctrl.passwords.get(from_partition_id, "")
        if from_pw:
            pw, ok = QInputDialog.getText(
                self, "验证来源分区密码",
                f"来源分区「{from_name}」设有密码，请输入密码：",
                QLineEdit.EchoMode.Password,
            )
            if not ok:
                return
            if pw.strip() != from_pw:
                QMessageBox.warning(self, "错误", "密码不正确")
                return

        # ── Step 2: Select target partition ──
        partitions = self._task_service.get_all_partitions()
        # Exclude current partition
        other = [p for p in partitions if p["id"] != from_partition_id]
        if not other:
            QMessageBox.information(self, "提示", "没有其他分区可供迁移。")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("选择目标分区")
        dlg.resize(320, 240)
        layout = QVBoxLayout(dlg)

        list_widget = QListWidget(dlg)
        for p in other:
            pid, pname = p["id"], p["name"]
            has_pw = bool(self._partition_ctrl.passwords.get(pid, ""))
            label = f"{'🔒 ' if has_pw else ''}{pname}"
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole, pid)
            list_widget.addItem(item)
        list_widget.setCurrentRow(0)
        layout.addWidget(list_widget)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, dlg
        )
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        selected = list_widget.currentItem()
        if selected is None:
            return
        to_partition_id = selected.data(Qt.ItemDataRole.UserRole)
        to_name = name_map.get(to_partition_id, to_partition_id)

        # ── Step 3: Verify TO partition password ──
        to_pw = self._partition_ctrl.passwords.get(to_partition_id, "")
        if to_pw:
            pw, ok = QInputDialog.getText(
                self, "验证目标分区密码",
                f"目标分区「{to_name}」设有密码，请输入密码：",
                QLineEdit.EchoMode.Password,
            )
            if not ok:
                return
            if pw.strip() != to_pw:
                QMessageBox.warning(self, "错误", "密码不正确")
                return

        # ── Step 4: Confirmation ──
        reply = QMessageBox.question(
            self, "确认操作",
            f"确认将 {len(ids)} 个任务从「{from_name}」移动到「{to_name}」？",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
        )
        if reply != QMessageBox.StandardButton.Ok:
            return

        # ── Step 5: Execute ──
        moved = self._task_service.batch_move_partition(ids, to_partition_id)
        if self._current_view == "edit":
            self._task_model.set_checked_ids(set())
            self._batch_toolbar.reset_toggle()
        else:
            self._refresh_batch_page()
        self._on_data_changed()
        self._flash_status(f"已将 {moved} 个任务移动至「{to_name}」")

    def _hide_confirm(self) -> None:
        self._confirm_bar.setVisible(False)
        self._batch_pending_action = {}

    # ------------------------------------------------------------------
    # Manual archive & clear
    # ------------------------------------------------------------------

    def _on_manual_archive(self) -> None:
        """Archive all DONE tasks in current partition (ignore archive_days threshold)."""
        pid = self._partition_ctrl.active_id
        if not pid:
            return
        f = TaskFilter()
        f.sort_by = self._filter_bar.build_filter().sort_by
        f.partition_id = pid
        f.statuses = {TaskStatus.DONE}
        done_tasks = self._task_service.search(f)
        if not done_tasks:
            QMessageBox.information(self, "归档", "当前分区没有已完成的任务。")
            return
        reply = QMessageBox.question(
            self, "确认归档",
            f"归档当前分区全部 {len(done_tasks)} 个已完成任务？",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
        )
        if reply == QMessageBox.StandardButton.Ok:
            ids = [t.id for t in done_tasks if not t.archived]
            if ids:
                self._task_service.archive_batch(ids)
                _log.info("Manual archive: %s tasks", len(ids))
            self._batch_page = 0
            self._refresh_batch_page()
            self._on_data_changed()
            self._flash_status(f"已归档 {len(ids)} 个任务")

    def _on_clear_archived(self) -> None:
        """Permanently delete all archived tasks in current partition."""
        pid = self._partition_ctrl.active_id
        if not pid:
            return
        f = TaskFilter()
        f.partition_id = pid
        f.show_archived = True
        all_tasks = self._task_service.search(f)
        archived = [t for t in all_tasks if t.archived]
        if not archived:
            QMessageBox.information(self, "清除", "当前分区没有已归档的任务。")
            return
        reply = QMessageBox.warning(
            self, "⚠ 确认清除",
            f"将永久删除 {len(archived)} 个已归档任务，此操作不可恢复！",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
        )
        if reply == QMessageBox.StandardButton.Ok:
            ids = [t.id for t in archived]
            self._task_service.batch_delete(ids)
            self._batch_page = 0
            self._refresh_batch_page()
            self._on_data_changed()
            self._flash_status(f"已清除 {len(ids)} 个已归档任务")

    # ------------------------------------------------------------------
    # Batch export
    # ------------------------------------------------------------------

    def _on_batch_export(self, fmt: str) -> None:
        """Export all tasks in current partition to MD or Excel."""
        pid = self._partition_ctrl.active_id
        f = TaskFilter()
        f.sort_by = self._filter_bar.build_filter().sort_by
        f.partition_id = pid
        tasks = self._task_service.search(f)
        if not tasks:
            QMessageBox.information(self, "导出", "当前分区没有任务。")
            return

        name_map = self._task_service.get_partition_name_map()
        pname = name_map.get(pid or "", "未知")
        today = date.today().isoformat()

        if fmt == "md":
            path, _ = QFileDialog.getSaveFileName(
                self, "导出 Markdown", f"{pname}_{today}.md",
                "Markdown 文件 (*.md);;所有文件 (*)"
            )
            if path:
                lines = [t.raw_md for t in tasks]
                with open(path, "w", encoding="utf-8") as fh:
                    fh.write("\n".join(lines) + "\n")
                self._flash_status(f"已导出 {len(tasks)} 个任务到 {path}")
        elif fmt == "xlsx":
            path, _ = QFileDialog.getSaveFileName(
                self, "导出 Excel", f"{pname}_{today}.xlsx",
                "Excel 文件 (*.xlsx);;所有文件 (*)"
            )
            if path:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = pname
                ws.append(["序号", "任务内容", "状态", "进度", "截止日期", "标签", "创建时间", "归档"])
                for i, t in enumerate(tasks, 1):
                    ws.append([
                        i, t.title, t.status.display_name, f"{t.progress}%",
                        t.deadline_date.isoformat() if t.deadline_date else "",
                        " ".join(f"#{tag}" for tag in t.tags),
                        t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                        "已归档" if t.archived else ("未归档" if t.status == TaskStatus.DONE else "/"),
                    ])
                wb.save(path)
                self._flash_status(f"已导出 {len(tasks)} 个任务到 {path}")

    def _on_filter_changed(self, filter_: TaskFilter) -> None:
        if self._setting_sort_internally:
            pass  # internal set_sort() call — keep new_task_sort_active
        elif self._new_task_sort_active:
            self._new_task_sort_active = False
            self._setting_sort_internally = True
            self._filter_bar.set_sort(self._config.default_sort)
            self._setting_sort_internally = False
            return  # set_sort triggers another _on_filter_changed with default
        self._carousel_filter = filter_
        self._refresh_all_views(filter_)

    def _on_quick_preset(self, preset: str) -> None:
        if self._new_task_sort_active:
            self._new_task_sort_active = False
            self._setting_sort_internally = True
            self._filter_bar.set_sort(self._config.default_sort)
            self._setting_sort_internally = False
        f = self._filter_bar.build_filter()  # preserve sort
        # 速览栏按创建时间过滤，排除已完成已归档（仓库层默认 archived=0 已排除已归档）
        if preset == "all":
            pass
        elif preset == "today":
            f.created_to = date.today()
        elif preset == "yesterday":
            f.created_to = date.today() - dt.timedelta(days=1)
        elif preset == "last_week":
            today = date.today()
            f.created_to = today - dt.timedelta(days=today.isoweekday())
        elif preset == "week":
            today = date.today()
            f.created_to = today + dt.timedelta(days=7 - today.isoweekday())
        elif preset == "last_month":
            today = date.today()
            f.created_to = today.replace(day=1) - dt.timedelta(days=1)
        elif preset == "month":
            import calendar as _cal
            today = date.today()
            _, last = _cal.monthrange(today.year, today.month)
            f.created_to = today.replace(day=last)
        elif "days" in preset:
            try:
                int(preset.split("_")[0])
                f.created_to = date.today()
            except ValueError:
                pass
        self._carousel_filter = f
        self._refresh_all_views(f)
        self._progress_bar.reset_to_unclicked()
        if hasattr(self, '_task_model') and self._task_model.rowCount() > 0:
            self._on_task_selected(self._task_model.tasks[0])
        if self._current_view != "edit":
            self._heatmap_widget.highlight_range(f.date_from, f.date_to, preset)

    def _on_status_clicked(self, status: TaskStatus) -> None:
        f = self._filter_bar.build_filter()
        f.statuses = [status] if status else None
        if self._carousel_filter:
            f.tags = list(self._carousel_filter.tags)
        self._carousel_filter = f
        self._refresh_all_views(f)

    def _on_progress_filter(self, filter_: TaskFilter) -> None:
        self._carousel_filter = filter_
        # 通过 _build_filter_with_sort() 合并速览栏 scope
        merged = self._build_filter_with_sort()
        self._refresh_all_views(merged)
        # 进度栏: 对任务列表额外做 activity_log 活动过滤
        if self._progress_bar._active_period:
            active_tasks = self._progress_bar.filter_tasks_by_activity(
                self._task_model.tasks
            )
            if len(active_tasks) < len(self._task_model.tasks):
                self._total_count = len(active_tasks)
                self._task_model.set_offset(0)
                self._task_model.load_tasks(active_tasks[:self._page_size])
                self._update_page_label()
        # 同步轮播数据到过滤后的列表，确保点击定位一致
        self._progress_bar.set_items(list(self._task_model.tasks))
        if self._task_model.tasks:
            self._on_task_selected(self._task_model.tasks[0])

    def _on_view_task_selected(self, task: Task) -> None:
        """Guard against signal recursion from selectRow() inside _on_task_selected."""
        if self._selection_guard:
            return
        self._on_task_selected(task)

    def _on_task_selected(self, task: Task) -> None:
        """统一任务凸显入口：模型凸显 + 编辑器加载 + 视图定位滚动。"""
        self._task_model.set_highlighted_task(task.id)
        self._edit_panel.load_task(task)
        self._last_activity = dt.datetime.now()
        # 在列表中定位并滚动到该任务
        for row in range(self._task_model.rowCount()):
            if self._task_model.tasks[row].id == task.id:
                self._selection_guard = True
                self._task_view.selectRow(row)
                self._task_view.scrollTo(self._task_model.index(row, 0))
                self._selection_guard = False
                break

    def _on_detail_requested(self, task: Task) -> None:
        self._edit_panel.load_task(task)

    def _on_task_selection_changed(self) -> None:
        selected = self._task_view.selected_task_ids()
        self._batch_toolbar.setVisible(len(selected) >= 1)

    def _on_model_data_changed(self) -> None:
        if hasattr(self, '_batch_toolbar'):
            ids = self._task_model.checked_task_ids()
            self._batch_toolbar.set_selected(ids)

    def _on_carousel_clicked(self, task_id: str) -> None:
        self._select_and_load_task(task_id)

    def _on_heatmap_data_changed(self, *args) -> None:
        if hasattr(self, '_heatmap_widget'):
            self._heatmap_widget.force_refresh()

    def _on_go_home(self) -> None:
        if self._current_view != "edit":
            self._switch_view("edit")
        self._filter_coordinator.go_home()

    def _on_escape(self) -> None:
        if self._current_view != "edit":
            self._switch_view("edit")
        else:
            self._on_go_home()

    # ------------------------------------------------------------------
    # Pagination
    # ------------------------------------------------------------------

    def _update_page_label(self) -> None:
        if self._page_size <= 0:
            self._page_label.setText("全部")
            return
        total_pages = max(1, (self._total_count + self._page_size - 1) // self._page_size)
        self._page_label.setText(f"{self._page + 1} / {total_pages}")

    def _on_page_prev(self) -> None:
        if self._page > 0:
            self._page -= 1
            self._refresh_all_views(self._build_filter_with_sort(), reset_page=False)
            if self._task_model.rowCount() > 0:
                self._on_task_selected(self._task_model.tasks[0])

    def _on_page_next(self) -> None:
        total_pages = max(1, (self._total_count + self._page_size - 1) // self._page_size)
        if self._page < total_pages - 1:
            self._page += 1
            self._refresh_all_views(self._build_filter_with_sort(), reset_page=False)
            if self._task_model.rowCount() > 0:
                self._on_task_selected(self._task_model.tasks[0])

    def _on_page_size_changed(self, index: int) -> None:
        widget = self.sender()
        if widget:
            self._page_size = widget.itemData(index)
            self._page = 0
            self._refresh_all_views(self._build_filter_with_sort(), reset_page=False)

    def _reset_pagination(self) -> None:
        self._page = 0

    # ------------------------------------------------------------------
    # Status bar helpers
    # ------------------------------------------------------------------

    def _update_status_bar(self, filter_: TaskFilter) -> None:
        counts = self._task_service.get_status_counts(partition_id=self._partition_ctrl.active_id)
        overdue = counts.get(TaskStatus.OVERDUE, 0)
        doing = counts.get(TaskStatus.DOING, 0)
        todo = counts.get(TaskStatus.TODO, 0)
        done = counts.get(TaskStatus.DONE, 0)
        total = sum(counts.values())
        preset = self._quick_overview._active_preset if hasattr(self._quick_overview, '_active_preset') else "all"
        breakdown = f"逾期 {overdue} | 进行中 {doing} | 待办 {todo} | 已完成 {done} | 共{total}项"
        self._status_msg.setText(breakdown)

    def _flash_status(self, msg: str) -> None:
        self._status_msg.setText(msg)
        QTimer.singleShot(3000, lambda: self._status_msg.setText("就绪"))

    # ------------------------------------------------------------------
    # Partition management
    # ------------------------------------------------------------------
    # Partition coordination — widget updates when partition changes
    # ------------------------------------------------------------------

    def _on_partition_activated(self, pid: str) -> None:
        """Update all partition-aware widgets after a partition is activated."""
        self._heatmap_widget.set_partition_id(pid or None)
        self._status_badge.set_partition_id(pid or None)
        self._progress_bar.set_partition_id(pid or None)
        self._quick_overview.set_partition_id(pid or None)
        self._filter_coordinator._progress_active = False
        self._filter_coordinator._carousel_filter = self._quick_overview.build_filter()
        self._filter_coordinator.set_partition(pid)
        self._filter_coordinator.refresh()
        self._batch_ctrl.refresh_page()
        self._batch_ctrl.set_active_partition(pid or None)
        self._heatmap_widget.force_refresh()
        if self._current_view == "dashboard":
            self._refresh_analysis(pid)
        if self._task_model.rowCount() > 0:
            self._filter_coordinator._on_task_selected(self._task_model.tasks[0])
        else:
            self._edit_panel.set_active_partition(pid)
            self._edit_panel.show_empty()

    def _on_partitions_changed(self) -> None:
        self._partition_ctrl.load_all()

    # ------------------------------------------------------------------
    # View switching
    # ------------------------------------------------------------------

    def _switch_view(self, view: str) -> None:
        if view == self._current_view:
            return
        _log.info("View switched: %s", view)
        self._current_view = view
        # Cancel any pending deferred loads
        if hasattr(self, '_deferred_timer') and self._deferred_timer.isActive():
            self._deferred_timer.stop()

        if view == "edit":
            self._stack.setCurrentIndex(0)
            self._heatmap_widget.nav_bar.setVisible(False)
            self._top_bar.show()
            self._apply_splitter_sizes()
        elif view == "dashboard":
            if not self._page1_built:
                self._build_page1()
            self._stack.setCurrentIndex(1)
            self._heatmap_widget.nav_bar.setVisible(True)
            self._top_bar.hide()
            self._deferred_timer = QTimer(self)
            self._deferred_timer.setSingleShot(True)
            self._deferred_timer.timeout.connect(
                lambda: self._refresh_analysis(self._partition_ctrl.active_id)
            )
            self._deferred_timer.start(0)
        elif view == "batch":
            if not self._batch_ctrl._built:
                self._build_page2()
            self._stack.setCurrentIndex(2)
            self._heatmap_widget.nav_bar.setVisible(False)
            self._top_bar.hide()
            self._apply_batch_splitter_sizes()
            self._batch_ctrl.refresh_page()
            self._batch_ctrl.set_active_partition(self._partition_ctrl.active_id or None)

        # Reset filter bar sort to config default on view switch
        self._new_task_sort_active = False
        self._filter_bar.set_sort(self._config.default_sort)

    def _load_dashboard_data(self) -> None:
        """Load dashboard stats after view switch (report loads on period click)."""
        self._refresh_analysis()

    # ------------------------------------------------------------------
    # Activity analysis slots
    # ------------------------------------------------------------------

    def _refresh_analysis(self, partition_id: str | None = None) -> None:
        """Refresh analysis page: heatmap stats + task tree."""
        if hasattr(self, '_analysis_stats') and hasattr(self, '_heatmap_widget'):
            model = self._heatmap_widget._model
            self._analysis_stats.refresh(
                total=model.total_count(),
                active_days=model.active_days(),
                longest_streak=model.longest_streak(),
                daily_avg=model.daily_average(),
            )
        d_from, d_to = getattr(self, '_analysis_date_range', (None, None))
        if hasattr(self, '_analysis_task_tree'):
            self._analysis_task_tree.refresh(d_from, d_to, partition_id)

    def _on_analysis_period_changed(self, d_from, d_to, label: str) -> None:
        """Period change → highlight heatmap + refresh task tree + select first task."""
        self._analysis_date_range = (d_from, d_to)
        if d_from is not None and d_to is not None:
            self._heatmap_widget.highlight_range(d_from, d_to, label)
        else:
            self._heatmap_widget.highlight_range(None, None, "")
        if hasattr(self, '_analysis_task_tree'):
            self._analysis_task_tree.refresh(d_from, d_to, self._partition_ctrl.active_id)

    def _on_analysis_tag_selected(self, tag: str) -> None:
        if not tag or not hasattr(self, '_analysis_content_view'):
            if hasattr(self, '_analysis_content_view'):
                self._analysis_content_view.show_hint()
            return
        tasks = self._analysis_task_tree.get_tasks_for_tag(tag)
        d_from, d_to = getattr(self, '_analysis_date_range', (None, None))
        checked = self._analysis_task_tree.get_checked_tags()
        if tag in checked:
            pos = checked.index(tag) + 1
        else:
            pos = 0
        self._analysis_content_view.set_current_tag(tag, pos, len(checked))
        self._analysis_content_view.show_tag_activity(tag, tasks, d_from, d_to)

    def _on_analysis_prev(self) -> None:
        if hasattr(self, '_analysis_task_tree'):
            self._analysis_task_tree.select_prev()

    def _on_analysis_next(self) -> None:
        if hasattr(self, '_analysis_task_tree'):
            self._analysis_task_tree.select_next()

    def _on_heatmap_date_clicked(self, d: date) -> None:
        """Handle date click on heatmap grid."""
        if hasattr(self, '_analysis_period_selector'):
            self._analysis_period_selector.set_custom_range(d, d)

    def _on_analysis_search_changed(self, text: str) -> None:
        """Filter activity content by search text."""
        if hasattr(self, '_analysis_content_view'):
            self._analysis_content_view.set_search_text(text)

    def _on_export_analysis_md(self) -> None:
        self._export_analysis("md")

    def _on_export_analysis_xlsx(self) -> None:
        self._export_analysis("xlsx")

    def _on_export_analysis_txt(self) -> None:
        self._export_analysis("txt")

    def _export_analysis(self, fmt: str) -> None:
        """Export analysis content for all checked tags to file."""
        d_from, d_to = getattr(self, '_analysis_date_range', (None, None))

        # Collect plain text from all checked tags
        texts: list[str] = []
        if hasattr(self, '_analysis_task_tree') and hasattr(self, '_analysis_content_view'):
            checked_tags = self._analysis_task_tree.get_checked_tags()
            for tag in checked_tags:
                tasks = self._analysis_task_tree.get_tasks_for_tag(tag)
                self._analysis_content_view.show_tag_activity(tag, tasks, d_from, d_to)
                t = self._analysis_content_view.get_plain_text()
                if t:
                    texts.append(t)
            # Restore current view
            current_tag = self._analysis_task_tree.get_active_tag()
            if current_tag:
                tasks = self._analysis_task_tree.get_tasks_for_tag(current_tag)
                self._analysis_content_view.show_tag_activity(current_tag, tasks, d_from, d_to)

        text = "\n".join(texts)
        if not text:
            return

        # Build default filename with tag count
        def_name = self._build_export_filename(fmt, len(checked_tags) if hasattr(self, '_analysis_task_tree') else 0)
        filters = {"md": "Markdown (*.md)", "xlsx": "Excel (*.xlsx)", "txt": "文本文件 (*.txt)"}
        filepath, _ = QFileDialog.getSaveFileName(self, "导出报告", def_name, filters.get(fmt, ""))
        if not filepath:
            return

        if fmt == "xlsx":
            self._export_xlsx_file(filepath, text)
        elif fmt == "md":
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(text)
        else:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(text)

    def _build_export_filename(self, fmt: str, tag_count: int = 0) -> str:
        """Build default export filename: 分区名_时间范围_n个标签.ext"""
        name_map = self._task_service.get_partition_name_map()
        pname = name_map.get(self._partition_ctrl.active_id or "", "默认分区")
        d_from, d_to = getattr(self, '_analysis_date_range', (None, None))
        date_str = ""
        if d_from and d_to:
            date_str = f"{d_from.isoformat()}" if d_from == d_to else f"{d_from.isoformat()}~{d_to.isoformat()}"
        tag_suffix = f"_{tag_count}个标签" if tag_count > 0 else ""
        return f"{pname}_{date_str}{tag_suffix}.{fmt}"

    def _export_xlsx_file(self, filepath: str, text: str = "") -> None:
        """Export as Excel with split columns: 序号, 任务, 状态变更, 进度变更, 活动信息."""
        try:
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "活动报告"
            headers = ["序号", "任务", "状态变更", "进度变更", "活动信息"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            rows = self._parse_export_rows(text)
            for r, row_data in enumerate(rows, 2):
                for c, val in enumerate(row_data, 1):
                    ws.cell(row=r, column=c, value=val)
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 60
            wb.save(filepath)
        except ImportError:
            QMessageBox.warning(self, "错误", "需要安装 openpyxl 库才能导出 Excel")

    def _parse_export_rows(self, text: str) -> list[tuple]:
        """Parse plain text format into Excel rows.
        Format:
            #tag
            1. title [status, prog]:
                entry line 1
                entry line 2
        """
        rows = []
        current_num = ""
        current_title = ""
        current_status = ""
        current_prog = ""
        current_entries: list[str] = []

        def _flush():
            if current_num and current_entries:
                rows.append((
                    int(current_num), current_title, current_status,
                    current_prog, "\n".join(current_entries)
                ))

        for line in text.split("\n"):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            # Ordered list item: "1. title [status, prog]:"
            if stripped[0].isdigit() and ". " in stripped:
                _flush()
                current_entries = []
                parts = stripped.split(". ", 1)
                current_num = parts[0]
                rest = parts[1]
                if " [" in rest and "]:" in rest:
                    current_title = rest.split(" [", 1)[0]
                    bracket = rest.split("[", 1)[1].split("]", 1)[0]
                    if ", " in bracket:
                        current_status, current_prog = bracket.split(", ", 1)
                    else:
                        current_status, current_prog = bracket, ""
                else:
                    current_title = rest.rstrip(":")
            elif line.startswith("    ") and current_num:
                current_entries.append(stripped)

        _flush()
        return rows

    # ------------------------------------------------------------------
    # Task operations
    # ------------------------------------------------------------------

    def _on_new_task(self) -> None:
        self._on_new_draft()

    def _on_menu_new_draft(self) -> None:
        """Menu bar: show window, discard any draft silently, create new."""
        if self._edit_panel.has_unsaved_draft():
            self._edit_panel.discard_draft()
        self._ensure_window_ready()
        self._edit_panel.create_draft_single()
        self._apply_splitter_sizes()

    def _on_menu_new_multi(self) -> None:
        """Menu bar: show window, discard any draft silently, create multi."""
        if self._edit_panel.has_unsaved_draft():
            self._edit_panel.discard_draft()
        self._ensure_window_ready()
        self._edit_panel.create_draft_multi()
        self._apply_splitter_sizes()

    def _ensure_window_ready(self) -> None:
        """Show, raise, and switch to edit view. Updates _current_view so
        subsequent view switches work correctly."""
        self._current_view = "edit"
        self._top_bar.show()
        self._heatmap_widget.nav_bar.setVisible(False)
        if self._splitter_stack.currentIndex() == 1:
            if self._partition_ctrl.passwords.get(self._partition_ctrl.active_id, ""):
                self._partition_ctrl.unlock()
            else:
                self._splitter_stack.setCurrentIndex(0)
        self._stack.setCurrentIndex(0)
        self.show()
        self.setWindowState(self.windowState() & ~Qt.WindowState.WindowMinimized)
        self.raise_()
        self.activateWindow()
        self._edit_panel.set_active_partition(self._partition_ctrl.active_id)

    def _on_new_draft(self) -> None:
        # From tray (window hidden): silently discard draft, no popup
        if not self.isVisible() and self._edit_panel.has_unsaved_draft():
            self._edit_panel.discard_draft()
        elif not self._guard_draft():
            return
        if self._splitter_stack.currentIndex() == 1:
            if self._partition_ctrl.passwords.get(self._partition_ctrl.active_id, ""):
                self._partition_ctrl.unlock()
                if self._splitter_stack.currentIndex() == 1:
                    return
            else:
                self._splitter_stack.setCurrentIndex(0)
        self._stack.setCurrentIndex(0)
        self.show()
        self.setWindowState(self.windowState() & ~Qt.WindowState.WindowMinimized)
        self.raise_()
        self.activateWindow()
        self._edit_panel.set_active_partition(self._partition_ctrl.active_id)
        self._edit_panel.create_draft()
        self._apply_splitter_sizes()

    def _guard_draft(self) -> bool:
        if not self._edit_panel.has_unsaved_draft():
            return True
        msg = QMessageBox(self)
        msg.setWindowTitle("未保存的草稿")
        msg.setText("当前有未保存的新建任务，是否保存？")
        save_btn = msg.addButton("保存", QMessageBox.ButtonRole.AcceptRole)
        discard_btn = msg.addButton("放弃", QMessageBox.ButtonRole.DestructiveRole)
        msg.addButton("取消", QMessageBox.ButtonRole.RejectRole)
        msg.exec()
        clicked = msg.clickedButton()
        if clicked == save_btn:
            self._edit_panel._on_save()
            return True
        if clicked == discard_btn:
            self._edit_panel.discard_draft()
            return True
        return False

    def _on_import(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "导入 Markdown", "", "Markdown 文件 (*.md *.txt);;所有文件 (*)"
        )
        if not path:
            return
        try:
            from ..services.md_importer import MarkdownImporter
            count = MarkdownImporter(self._repository).import_file(path)
            _log.info("Import: %s -> %s tasks", path, count)
            self._on_data_changed()
            self._flash_status(f"已导入 {count} 个任务")
        except Exception as e:
            _log.error("Import failed: %s", e)
            QMessageBox.warning(self, "导入失败", str(e))

    def _on_export(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Markdown", "tasks.md", "Markdown 文件 (*.md);;所有文件 (*)"
        )
        if not path:
            return
        try:
            from ..services.md_exporter import MarkdownExporter
            tasks = self._task_service.get_all()
            MarkdownExporter.export_to_file(tasks, path)
            _log.info("Export: %s -> %s tasks", path, len(tasks))
            self._flash_status(f"已导出到 {path}")
        except Exception as e:
            _log.error("Export failed: %s", e)
            QMessageBox.warning(self, "导出失败", str(e))

    def _on_settings(self) -> None:
        _before = json.dumps(self._config.to_dict(), sort_keys=True)
        dlg = SettingsDialog(
            self._config, self._repository,
            task_service=self._task_service, parent=self,
        )
        if dlg.exec() == SettingsDialog.DialogCode.Accepted:
            _after = json.dumps(self._config.to_dict(), sort_keys=True)
            if _before == _after:
                return  # 零变更，跳过全部刷新
            # 设置保存后强制激活默认分区
            self._partition_ctrl.load_all()
            self._signal_bus.config_changed.emit()

    def _on_about(self) -> None:
        dlg = AboutDialog(self, update_checker=self._update_checker)
        dlg.exec()

    def _on_help_docs(self) -> None:
        import sys
        from pathlib import Path

        from PySide6.QtCore import QUrl
        from PySide6.QtGui import QDesktopServices

        base = getattr(sys, "_MEIPASS", None)
        if base:
            path = Path(base) / "resources" / "help" / "manual.html"
        else:
            path = Path(__file__).resolve().parents[2] / "resources" / "help" / "manual.html"
        if path.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))

    def _on_quit(self) -> None:
        self._signal_bus.application_quit.emit()

    def _on_refresh(self) -> None:
        self._filter_coordinator.refresh()

    def _on_config_changed(self) -> None:
        data_changed = False
        theme_changed = self._config.theme != self._last_applied_theme

        self._filter_bar.set_sort(self._config.default_sort)

        # Only refresh theme-dependent widgets when the theme actually changed
        if theme_changed:
            if hasattr(self, '_status_badge'):
                self._status_badge.refresh_theme()
            tag_panel = self._batch_ctrl.tag_panel
            if tag_panel is not None:
                tag_panel.refresh_theme()
            self._last_applied_theme = self._config.theme

        # Re-read page_size from config and sync all pagination controls
        new_page_size = self._config.get("general", "page_size", default=20)
        if self._page_size != new_page_size:
            self._page_size = new_page_size
            self._page = 0
            data_changed = True
            if hasattr(self, '_page_size_combo'):
                self._page_size_combo.setCurrentText(str(new_page_size))
        if hasattr(self, '_batch_page_size') and self._batch_page_size != new_page_size:
            self._batch_page_size = new_page_size
            self._batch_page = 0
            data_changed = True
            if hasattr(self, '_batch_page_size_combo'):
                self._batch_page_size_combo.setCurrentText(str(new_page_size))
        tag_panel = self._batch_ctrl.tag_panel
        if tag_panel is not None:
            tag_panel.set_page_size(new_page_size)

        # Heatmap: repaint on colour-scheme change (refresh_tokens already called by app.py)
        if hasattr(self, '_heatmap_widget'):
            self._heatmap_widget.force_refresh()

        # Sync completed-last sort setting to repository
        if self._task_service._repo.completed_last != self._config.sort_completed_last:
            self._task_service._repo.completed_last = self._config.sort_completed_last
            data_changed = True

        # Only reload task data when sort/pagination actually changed
        if data_changed:
            self._on_data_changed()

        # If default partition changed in settings, sync status bar
        default_pid = self._config.get("general", "default_partition", default="")
        current_pid = self._partition_ctrl.active_id or ""
        if default_pid and default_pid != current_pid:
            self._partition_ctrl.activate(default_pid)

    # ------------------------------------------------------------------
    # Midnight timer
    # ------------------------------------------------------------------

    def _setup_midnight_timer(self) -> None:
        self._midnight_timer = QTimer(self)
        self._midnight_timer.setSingleShot(True)
        self._midnight_timer.timeout.connect(self._on_midnight_crossed)
        self._schedule_midnight_timer()

    def _schedule_midnight_timer(self) -> None:
        now = QDateTime.currentDateTime()
        tomorrow = now.addDays(1)
        midnight = QDateTime(tomorrow.date(), QTime(0, 0, 1))
        ms = now.msecsTo(midnight)
        if ms <= 0:
            ms = 1000
        self._midnight_timer.start(ms)

    def _on_midnight_crossed(self) -> None:
        self._quick_overview.refresh()
        if self._current_view != "edit":
            self._refresh_report()
        self._on_data_changed()
        self._schedule_midnight_timer()
