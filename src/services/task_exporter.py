"""Excel task exporter — shared by the batch UI and the CLI."""

from __future__ import annotations

from ..models.task import Task

_COLUMNS = ["ID", "标题", "状态", "标签", "截止日", "进度"]


def export_xlsx(tasks: list[Task], path: str) -> int:
    """Write tasks to an ``.xlsx`` file at ``path``. Returns task count."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务列表"
    for col, header in enumerate(_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    for r, task in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=task.id)
        ws.cell(row=r, column=2, value=task.title)
        ws.cell(row=r, column=3, value=task.status.display_name)
        ws.cell(row=r, column=4, value=", ".join(task.tags))
        ws.cell(row=r, column=5, value=task.deadline_date.isoformat() if task.deadline_date else "")
        ws.cell(row=r, column=6, value=task.progress)
    wb.save(path)
    return len(tasks)
